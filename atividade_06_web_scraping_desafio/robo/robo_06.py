from selenium import webdriver
from selenium.webdriver.common.by import By
import pyautogui as p 
from time import sleep
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
import os
import sys

def excluir_cookies(xpath):
    cookie = xpath
    driver.find_element(By.XPATH,cookie).click()
    

def excluir_propaganda(xpath):
    prop = xpath
    driver.find_element(By.XPATH,prop).click()


def abrir_e_preparar_site(driver,url):
    driver.get(url)
    driver.maximize_window()
    p.press('F11')
    sleep(7)
    xpath_cookies = '/html/body/div[3]/div[1]/div/div/div[2]/button[2]/span'
    xpath_propagandas = '//*[@id="clever_56734_close_btn"]'
    excluir_cookies(xpath_cookies)
    excluir_propaganda(xpath_propagandas)
    

def pega_dados_e_retorna_em_variaveis(qtde):
    #Xpaths do porte estavam dando vários problemas, por isso guardei-os em um array para serem utilizados de maneira mais eficiente
    xpath_porte = ['/html/body/div[1]/div[1]/div/div[1]/p[4]/b','/html/body/div/div[1]/div/div[1]/p[5]/b','/html/body/div/div[1]/div/div[1]/p[4]/b']
    pessoa = int(qtde)
    #clica na pessoa
    driver.find_element(By.CSS_SELECTOR ,f"body > div.bg-white > main > div.max-w-4xl.mx-auto.bg-white.shadow.overflow-hidden.sm\:rounded-md > ul > li:nth-child({pessoa}) > a > div > div.flex.items-center.justify-between > p").click()
    sleep(2)
    #Nome:
    nome = driver.find_element(By.XPATH,'/html/body/div/div[1]/div/div[1]/p[2]/span/b').text
    #dias:
    dias = driver.find_element(By.XPATH,'/html/body/div[1]/div[1]/div/div[1]/p[3]/span[1]/b').text                              
    if dias[0:2].isdigit():
        dias = dias  
    else:
        dias = driver.find_element(By.XPATH,'/html/body/div/div[1]/div/div[1]/p[4]/span[1]/b').text
    #porte
    try:
        porte = driver.find_element(By.XPATH,xpath_porte[0]).text
    except:
        sleep(0.5)
    try:       
        porte = driver.find_element(By.XPATH,xpath_porte[1]).text
    except:
        sleep(0.5)
    try:       
        porte = driver.find_element(By.XPATH,xpath_porte[2]).text
    except:
        sleep(0.5)
    return nome,dias,porte


def armazenar_informacoes_em_listas():
    qtde = 1
    nome_list = []
    dias_list = []
    porte_list = []
    sit_list = []
    for l in range(1):
        qtde = 1
        sleep(1.5)
        while qtde >=1 and qtde<28:
            print(f'raspando dados da pagina {l + 1} empresa {qtde}')
            if qtde == 14 or qtde == 15:
                qtde += 1
            else:
                situacao = driver.find_element(By.CSS_SELECTOR ,f"body > div > main > div.max-w-4xl.mx-auto.bg-white.shadow.overflow-hidden.sm\:rounded-md > ul > li:nth-child({qtde}) > a > div > div.flex.items-center.justify-between > div").text
                if situacao.upper() == 'ATIVA':
                    nome,dias,porte = pega_dados_e_retorna_em_variaveis(qtde)
                    nome_list.append(nome)
                    dias_list.append(dias)
                    porte_list.append(porte)
                    sit_list.append(situacao)
                    p.hotkey('browserback')
                    qtde += 1
                else:
                    nome,dias,porte = pega_dados_e_retorna_em_variaveis(qtde)
                    nome_list.append(nome)
                    dias_list.append('')
                    porte_list.append('')
                    sit_list.append(situacao)
                    p.hotkey('browserback')
                    qtde += 1
        sleep(0.5)
        driver.find_element(By.XPATH,'/html/body/div[1]/main/div[3]/div[1]/div/a').click()
        print(f'\nCOMPLETOU A PÁGINA {l + 1}\n')
        sleep(1)

    return nome_list,dias_list,porte_list,sit_list   


def salva_em_excell_e_altera_cor_dos_inativos(lista_1,lista_2,lista_3,lista_4):
    #abre o arquivo
    book = openpyxl.Workbook('Dados_empresa.xlsx')
    #cria a página
    book.create_sheet('dados')
    #adiciona a página à uma váriavel
    dados_page = book['dados']
    #adicona os primeiros campos
    dados_page.append(['NOME','DATA_ABERTURA','PORTE','SITUAÇÃO'])
    #adiciona e separa dentro do excel qual está ativa ou nao
    for x in range(len(lista_1)-1):
        if (lista_2[x] != "red"):
            dados_page.append([f'{lista_1[x]}',f'{lista_2[x]}',f'{lista_3[x]}',f'{lista_4[x]}'])
        else:
            dados_page.append([f'{lista_1[x]}','','',f'{lista_4[x]}'])
    #salvando informacoes na tabela
    book.save('Dados_empresa.xlsx')
    #alterando tamanho da tabela
    wb = load_workbook('Dados_empresa.xlsx')
    ws = wb['dados']
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 16
    My_font = Font(color='FF0000')
    #alterando as cores dos campos que não são ativos
    for rows in ws.iter_rows(2):
        if rows[3].value != 'ATIVA':
            rows[0].font = My_font
            rows[1].font = My_font
            rows[2].font = My_font
            rows[3].font = My_font
    #salva Definitivamente
    wb.save('Dados_empresa.xlsx')
    #fecha a planilha
    wb.close()
    
def bot_006(driver):
    try:
        list1 = []
        list2 = []
        list3 = []
        abrir_e_preparar_site(driver,'https://cnpj.biz/empresas/blumenau-sc')
        sleep(1)
        list1,list2,list3,list4 = armazenar_informacoes_em_listas()
        salva_em_excell_e_altera_cor_dos_inativos(list1,list2,list3,list4)
        p.press('F11')
        driver.close()     
    except:
        exc_type, error, line = sys.exc_info()
        print(f'ERROR: {error}\nCLASS: {exc_type}\nFUNC: {sys._getframe().f_code.co_name}\nLINE:  {line.tb_lineno}\n')


if (__name__ == '__main__'):
#    laziest solution
    os.system('cls')
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = webdriver.Chrome(options=options)
    bot_006(driver)
    print('Robô executado com sucesso')

''''
bro one question:

im doing the excel thing, from our robot. and the informations comes in that whay 



do you know how to "expand" the cell? or one bether whay to make this informations be clear then now?



'''