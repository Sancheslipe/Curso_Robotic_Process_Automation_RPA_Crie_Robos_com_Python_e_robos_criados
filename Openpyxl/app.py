import openpyxl

#criar planilha 
book = openpyxl.Workbook()
#visualizar paginas existentes
print(book.sheetnames)
#criar uma pagina
book.create_sheet('Frutas')
#Como selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Fruta','Quantidade','Preço'])
frutas_page.append(['banana','5','R$3,90'])
frutas_page.append(['banana2','5','R$3,90'])
frutas_page.append(['banana3','5','R$3,90'])
frutas_page.append(['banana4','5','R$3,90'])
#salvarplanilha
book.save('Planilha de compras.xlsx')
