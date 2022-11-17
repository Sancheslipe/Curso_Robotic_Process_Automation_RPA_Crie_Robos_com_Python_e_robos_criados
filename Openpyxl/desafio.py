import openpyxl

book = openpyxl.Workbook('meus computadores.xlsx')
wb = openpyxl.Workbook()
book.create_sheet('computadores')
note_page = book['computadores']
note_page.append(['Eletrônica','memória ram','preço'])
soma = 0
cont = 0 
while cont < 1:
    qtde = input('digite aqui a quantidade de computadores que deseja adicionar no arquivo: ')
    if qtde.isdigit():
        qtde = int(qtde)
        cont += 1 
    else:
        cont = cont
        print('digite um numero válido')
#adicionando manualmente
# note_page.append(['Computador 1',' 8gb ram','R$2500'])

cont = 0
while cont < qtde:
    computador = input('\ndigite aqui o nome do computador: ')
    ram = input('digite aqui a quantidade de memoria ram: ')
    if ram.isdigit():
        ram = int(ram)
        valor = input('digite aqui o valor do computador: ')
        if valor.replace('.','', 1).isdigit():
            valor = float(valor)
            soma += valor
            note_page.append([f'{computador}',f'{ram}gb ram',f'{valor}'])
            cont += 1
        else:
            cont = cont
            print('digite um valor válido para o preço do notebook! \n')
    else:
        cont = cont
        print('digite um numero válido para memoria ram!\n') 
note_page.append(['Valor Total:',soma])
book.save('meus computadores.xlsx')
print('salvou')

# def formata_arq(dir_path, f):
#     try:
#         wb = openpyxl.load_workbook(f"{dir_path}\\{f}")
#         ws = wb['Sheet1']

#         for row in ws.iter_rows():
#             for cell in row:
#                 if cell.value == '*':
#                     ws[cell.coordinate].font = Font(color='FF0000')

#         wb.save(f"{dir_path}\\{f}")
#     except Exception as e:
#         raise Exception(f'Erro {e} em formata_arq')